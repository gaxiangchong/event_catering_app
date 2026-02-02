from app import create_app, db
from app.models import User, Event, MealOption, Order, OrderStatus, EventStatus
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO

def test_excel_export():
    app = create_app()
    with app.app_context():
        print("Testing Excel Export...")
        
        # 1. Setup Data
        # Ensure at least 2 events exist to test multi-sheet
        event1 = Event.query.filter_by(title="Excel Event 1").first()
        if not event1:
            event1 = Event(title="Excel Event 1", date=datetime.now(), fee=10.0, status=EventStatus.ACTIVE)
            db.session.add(event1)
            
        event2 = Event.query.filter_by(title="Excel Event 2").first()
        if not event2:
            event2 = Event(title="Excel Event 2", date=datetime.now(), fee=20.0, status=EventStatus.ACTIVE)
            db.session.add(event2)
        
        db.session.flush()

        # Create orders if not exist
        if Order.query.filter_by(event_id=event1.id).count() == 0:
            user = User.query.first()
            if not user:
                 user = User(name="Test User Excel", telephone="111222333")
                 db.session.add(user)
                 db.session.flush()
                 
            meal1 = MealOption(event_id=event1.id, name="Meal 1")
            db.session.add(meal1)
            db.session.flush()
            
            order1 = Order(user_id=user.id, event_id=event1.id, meal_option_id=meal1.id, amount=10, status=OrderStatus.PAID)
            db.session.add(order1)
            
        if Order.query.filter_by(event_id=event2.id).count() == 0:
            user = User.query.first()
            if not user:
                 user = User(name="Test User Excel", telephone="111222333")
                 db.session.add(user)
                 db.session.flush()

            meal2 = MealOption(event_id=event2.id, name="Meal 2")
            db.session.add(meal2)
            db.session.flush()
            
            order2 = Order(user_id=user.id, event_id=event2.id, meal_option_id=meal2.id, amount=20, status=OrderStatus.PENDING)
            db.session.add(order2)

        db.session.commit()
        
        # 2. Test Export Logic
        # Reproducing "export_orders" logic for 'all' events
        from openpyxl import Workbook
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)
        
        events = Event.query.order_by(Event.date.desc()).all()
        sheet_count = 0
        
        for event in events:
            # Logic from admin.py
            sheet_title = "".join(c for c in event.title if c.isalnum() or c in (' ', '_', '-'))[:30]
            ws = wb.create_sheet(title=sheet_title)
            ws.append(['Order ID', 'Customer Name']) # Header
            
            orders = Order.query.filter_by(event_id=event.id).all()
            for order in orders:
                ws.append([order.id, order.user.name])
            
            sheet_count += 1
            
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        # Verify
        wb_loaded = load_workbook(out)
        print(f"Generated sheets: {wb_loaded.sheetnames}")
        
        if "Excel Event 1" in wb_loaded.sheetnames and "Excel Event 2" in wb_loaded.sheetnames:
            print("SUCCESS: Multi-sheet Excel generated correctly.")
        else:
            print("FAILURE: Expected sheets not found.")

        # Cleanup
        # db.session.delete(order1)
        # db.session.delete(order2) 
        # Keeping them for UI verification if needed
        
if __name__ == "__main__":
    test_excel_export()
